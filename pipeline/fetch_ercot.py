"""Fetch + parse the ERCOT GIS Report (generation interconnection queue).

Downloads the latest monthly GIS_Report xlsx from ERCOT MIS (reportTypeId 15933)
and extracts active queue projects (Large Gen + Small Gen tabs) with county,
POI, MW, fuel, study phase, and projected COD.
"""
import re
import sys

import openpyxl

sys.path.insert(0, "pipeline")
from common import RAW, http_bytes, http_json, save_json

LIST_URL = "https://www.ercot.com/misapp/servlets/IceDocListJsonWS?reportTypeId=15933"
DL_URL = "https://www.ercot.com/misdownload/servlets/mirDownload?doclookupId={}"

INR_RE = re.compile(r"^\d{2}INR\d+", re.I)


def download_latest():
    docs = http_json(LIST_URL)["ListDocsByRptTypeRes"]["DocumentList"]
    gis = [d["Document"] for d in docs if d["Document"]["FriendlyName"].startswith("GIS_Report")]
    gis.sort(key=lambda d: d["PublishDate"], reverse=True)
    latest = gis[0]
    print(f"latest: {latest['FriendlyName']} published {latest['PublishDate']}")
    data = http_bytes(DL_URL.format(latest["DocID"]))
    path = f"{RAW}/GIS_Report.xlsx"
    with open(path, "wb") as f:
        f.write(data)
    return path, latest["FriendlyName"]


def parse_sheet(ws, sheet_label):
    rows = list(ws.iter_rows(values_only=True))
    header_i = None
    for i, row in enumerate(rows):
        if row and str(row[0]).strip() == "INR":
            header_i = i
            break
    if header_i is None:
        return []
    out = []
    for row in rows[header_i + 1:]:
        if not row or not row[0] or not INR_RE.match(str(row[0]).strip()):
            continue
        def cell(j):
            v = row[j] if j < len(row) else None
            return str(v).strip() if v is not None else ""
        mw = row[10] if len(row) > 10 else None
        try:
            mw = float(mw)
        except (TypeError, ValueError):
            mw = None
        out.append({
            "inr": cell(0), "name": cell(1), "phase": cell(2),
            "entity": cell(3), "poi": cell(4), "county": cell(5),
            "zone": cell(6), "cod": cell(7)[:10], "fuel": cell(8),
            "tech": cell(9), "mw": mw, "sheet": sheet_label,
        })
    return out


def main():
    path, name = download_latest()
    wb = openpyxl.load_workbook(path, read_only=True)
    projects = []
    for sheet, label in [("Project Details - Large Gen", "large"),
                         ("Project Details - Small Gen", "small")]:
        if sheet in wb.sheetnames:
            got = parse_sheet(wb[sheet], label)
            print(f"{sheet}: {len(got)} projects")
            projects.extend(got)
    save_json(f"{RAW}/ercot_queue.json", {"report": name, "projects": projects})


if __name__ == "__main__":
    main()
